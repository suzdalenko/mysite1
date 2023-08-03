from openpyxl.styles import Alignment, PatternFill, Font, Color
import openpyxl
from myapp.models import CollectionLines
from myapp.myresponse import SuzdalenkoJsonResponse, OrderingPackagesByTruck

class Api:

    def test_one(request):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        ws.merged_cells.ranges.add("A1:E1")
        
        ws['A1'] = "RUTAS DE CAMIONES"
        ws.append(["xxxxxxxxxxxx", "xxxxxxxxxxxx", "xxxxxxxxxxxx"])
        ws.append(["xxxxxxxxxxxx", "xxxxxxxxxxxx", "xxxxxxxxxxxx"])

        ws['A2'] = "suzdalenko alexey"

        a1 = ws['A1']
        ft = Font(color="FFFFFF", italic=True, size=22)
        a1.font = ft
        a1.alignment = Alignment(horizontal='center')
        a1.fill = PatternFill(start_color="4669aa", end_color="4669aa", fill_type = "solid")

        ws.column_dimensions['A'].width = 11

        dims = {}
        letter = ['A', 'B', 'C', 'D']
        for row in ws.rows:
            for cell in row:
                if cell.value:
                  dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
        for col, value in dims.items():
            print(letter[col])
            ws.column_dimensions[letter[col]].width = value + 3

        wb.save("sample.xlsx")
    
    #   wb = openpyxl.Workbook()
    #   # ws = wb.create_sheet("Test")
    #   ws = wb.active
    #   ws['A1'] = "xxx"

    #   # ws.append(['A', '', '', '', '', 'B', '', '', '', ''])
    #   # ws.append([1, 2, 33, '', '', 'D', '', '', '', ''])
# 
    #   # ws.merged_cells.ranges.add("A1:E1")
    #   # ws.merged_cells.ranges.add("F1:J1")
# 
    #   # currentCell = ws.cell('A1')
    #   # currentCell.alignment = Alignment(horizontal='center')
    #   
    #   # a1 = ws['A1']
    #   # a1.font = Font(color="FF0000", italic=True) 
    #   

    #   wb.save("x.xlsx")



        return SuzdalenkoJsonResponse({"route":"deleted"}) 